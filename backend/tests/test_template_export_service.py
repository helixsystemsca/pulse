"""Material requisition template export — dropdown validations and header fields."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.services.template_export_paths import backend_root, templates_dir
from app.services.template_export_service import (
    TemplateExportService,
    build_material_request_template_form,
    read_workbook_list_range,
    load_template,
    load_template_map,
)

TEMPLATES = Path(__file__).resolve().parents[2] / "templates"


@pytest.fixture
def export_service() -> TemplateExportService:
    return TemplateExportService(TEMPLATES / "template-map.json")


def _validation_map(ws) -> dict[str, str]:
    out: dict[str, str] = {}
    for dv in ws.data_validations.dataValidation:
        formula = (dv.formula1 or "").lstrip("=")
        for ref in str(dv.sqref).split():
            out[ref] = formula
    return out


def test_export_preserves_line_item_dropdown_validations(export_service: TemplateExportService) -> None:
    rows = [
        {
            "item_name": "Dish Soap",
            "vendor": "Vallen",
            "quantity": 2,
            "unit": "EACH",
            "reimbursable": False,
        }
    ]
    data, _ = export_service.export(rows, project="KEARL", location="Office", requester_name="Jane Doe")
    wb = load_workbook(BytesIO(data))
    ws = wb["MR Tracker"]
    validations = _validation_map(ws)

    assert len(validations) >= 3
    assert validations.get("B5:B23") == "Lists!$A$4:$A$5"
    assert validations.get("C5:C23") == "Lists!$C$4:$C$6"
    assert validations.get("I5:I23") == "Lists!$F$4:$F$19"


def test_export_preserves_requester_and_site_manager_name_dropdowns(
    export_service: TemplateExportService,
) -> None:
    rows = [{"item_name": "Item", "vendor": "V", "quantity": 1, "unit": "EACH", "reimbursable": True}]
    data, _ = export_service.export(rows, project="KEARL", location="Office", requester_name="Jane Doe")
    wb = load_workbook(BytesIO(data))
    ws = wb["MR Tracker"]
    validations = _validation_map(ws)

    assert validations.get("C24:D24") == "Lists!$H$4:$H$19"
    assert validations.get("C26:D26") == "Lists!$J$4:$J$5"
    assert ws["C24"].value == "Jane Doe"


def test_template_form_loads_modal_dropdown_options() -> None:
    form = build_material_request_template_form(load_template_map(TEMPLATES / "template-map.json"))
    by_key = {f["key"]: f for f in form["fields"]}
    assert by_key["cost_object"]["options"] == [
        "4503048718 - I&SS",
        "4503048145 - Mine",
        "OTHER - (specify in the comments)",
    ]
    assert by_key["location"]["options"] == [
        "Plant / Infrastructure",
        "Mine OPS & AHS - General",
        "OTHER",
    ]
    assert by_key["project"]["options"] == []


def test_templates_resolve_from_backend_bundle() -> None:
    bundled = backend_root() / "templates" / "template-map.json"
    assert bundled.is_file(), "backend/templates must ship with the API for Render deploys"
    assert templates_dir().is_dir()
    assert (templates_dir() / "kent_material_request.xlsx").is_file()


def test_read_workbook_list_range_parses_sheet_reference() -> None:
    tm = load_template_map(TEMPLATES / "template-map.json")
    wb = load_template(tm)
    assert read_workbook_list_range(wb, "Lists!$A$4:$A$5") == ["YES", "NO"]


def test_export_shifts_footer_validations_when_extra_line_rows_are_inserted(
    export_service: TemplateExportService,
) -> None:
    rows = [
        {
            "item_name": f"Item {i}",
            "vendor": "V",
            "quantity": 1,
            "unit": "EACH",
            "reimbursable": i % 2 == 0,
        }
        for i in range(22)
    ]
    data, _ = export_service.export(rows, project="KEARL", location="Office", requester_name="Jane Doe")
    wb = load_workbook(BytesIO(data))
    ws = wb["MR Tracker"]
    validations = _validation_map(ws)

    assert validations.get("B5:B26") == "Lists!$A$4:$A$5"
    assert validations.get("C27:D27") == "Lists!$H$4:$H$19"
    assert validations.get("C29:D29") == "Lists!$J$4:$J$5"
    assert ws["A27"].value == "Requester:"
