"""Excel export for material request drafts."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

from app.models.domain import MaterialRequestDraft, MaterialRequestDraftItem


def build_material_request_workbook(
    draft: MaterialRequestDraft,
    items: list[MaterialRequestDraftItem],
) -> tuple[bytes, str]:
    wb = Workbook()
    ws = wb.active
    ws.title = "Material Request"

    headers = [
        "Description",
        "Part Number",
        "Vendor",
        "Quantity",
        "Unit Cost",
        "Extended Cost",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for it in items:
        qty = float(it.qty_requested)
        unit = float(it.estimated_unit_cost) if it.estimated_unit_cost is not None else None
        extended = float(it.estimated_cost) if it.estimated_cost is not None else (
            round(qty * unit, 2) if unit is not None else None
        )
        ws.append(
            [
                it.item_name,
                it.sku,
                it.vendor or "",
                qty,
                unit if unit is not None else "",
                extended if extended is not None else "",
            ]
        )

    ws.append([])
    total_row = ["", "", "", "", "Estimated total", ""]
    total = 0.0
    for it in items:
        if it.estimated_cost is not None:
            total += float(it.estimated_cost)
        elif it.estimated_unit_cost is not None:
            total += float(it.qty_requested) * float(it.estimated_unit_cost)
    total_row[5] = round(total, 2)
    ws.append(total_row)
    ws[f"E{ws.max_row}"].font = Font(bold=True)
    ws[f"F{ws.max_row}"].font = Font(bold=True)

    buf = BytesIO()
    wb.save(buf)
    filename = f"{draft.draft_number}.xlsx"
    return buf.getvalue(), filename
