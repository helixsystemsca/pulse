"""Reusable Excel template export engine (load map, populate, preserve styles)."""

from __future__ import annotations

import json
import re
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from app.services.template_export_paths import template_map_path, template_workbook_path


class TemplateExportError(Exception):
    """User-facing export failure."""


@dataclass(frozen=True)
class TemplateMap:
    raw: dict[str, Any]

    @property
    def template_file(self) -> str:
        return str(self.raw.get("templateFile") or "kent_material_request.xlsx")

    @property
    def sheet_name(self) -> str | None:
        name = self.raw.get("sheetName")
        return str(name) if name else None

    @property
    def first_data_row(self) -> int:
        return int(self.raw.get("firstDataRow") or 8)

    @property
    def last_template_data_row(self) -> int:
        return int(self.raw.get("lastTemplateDataRow") or self.first_data_row)

    def header_cell(self, key: str) -> str | None:
        direct = self.raw.get(f"{key}Cell")
        if direct:
            return str(direct)
        fields = self.raw.get("fields") or {}
        block = fields.get(key) or {}
        cell = block.get("cell")
        return str(cell) if cell else None

    def populate_columns(self) -> dict[str, str]:
        cols = self.raw.get("populateColumns") or self.raw.get("columns") or {}
        return {str(k): str(v) for k, v in cols.items()}

    def data_validation_config(self) -> dict[str, Any]:
        block = self.raw.get("dataValidations")
        return dict(block) if isinstance(block, dict) else {}

    def modal_fields(self) -> list[dict[str, Any]]:
        raw = self.raw.get("modalFields")
        if not isinstance(raw, list):
            return []
        return [dict(item) for item in raw if isinstance(item, dict)]


def load_template_map(path: Path | None = None) -> TemplateMap:
    map_path = path or template_map_path()
    if not map_path.is_file():
        raise TemplateExportError(f"Template mapping file not found: {map_path}")
    with map_path.open(encoding="utf-8") as f:
        return TemplateMap(json.load(f))


def _parse_excel_range_ref(range_ref: str) -> tuple[str, int, int, int, int]:
    ref = (range_ref or "").strip()
    if not ref:
        raise TemplateExportError("Empty list range")
    if "!" in ref:
        sheet_part, cell_part = ref.split("!", 1)
        sheet = sheet_part.strip().strip("'").strip('"')
    else:
        sheet = ""
        cell_part = ref
    cell_part = cell_part.replace("$", "")
    min_col, min_row, max_col, max_row = range_boundaries(cell_part)
    if not sheet:
        raise TemplateExportError(f"List range must include a sheet name: {range_ref}")
    return sheet, min_col, min_row, max_col, max_row


def read_workbook_list_range(wb: Any, range_ref: str) -> list[str]:
    """Read non-empty unique cell values from an Excel range (e.g. Lists!$C$4:$C$6)."""
    sheet, min_col, min_row, max_col, max_row = _parse_excel_range_ref(range_ref)
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    values: list[str] = []
    seen: set[str] = set()
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            raw = ws.cell(row, col).value
            if raw is None:
                continue
            text = str(raw).strip()
            if not text or text in seen:
                continue
            seen.add(text)
            values.append(text)
    return values


def build_material_request_template_form(template_map: TemplateMap | None = None) -> dict[str, Any]:
    """Form metadata for MR export modals — options loaded from the template Lists sheet."""
    tm = template_map or load_template_map()
    fields_cfg = tm.modal_fields()
    if not fields_cfg:
        fields_cfg = [
            {"key": "project", "label": "Project", "required": True, "placeholder": "e.g. KEARL"},
            {"key": "cost_object", "label": "Cost object", "required": False, "placeholder": "Optional"},
            {
                "key": "location",
                "label": "Job description / location",
                "required": True,
                "placeholder": "e.g. Office consumables",
            },
            {"key": "comments", "label": "Comments", "required": False, "multiline": True},
        ]

    wb = load_template(tm)
    out_fields: list[dict[str, Any]] = []
    for item in fields_cfg:
        key = str(item.get("key") or "").strip()
        if not key:
            continue
        options: list[str] = []
        list_range = str(item.get("listRange") or "").strip()
        if list_range:
            try:
                options = read_workbook_list_range(wb, list_range)
            except TemplateExportError:
                options = []
        out_fields.append(
            {
                "key": key,
                "label": str(item.get("label") or key.replace("_", " ").title()),
                "required": bool(item.get("required")),
                "placeholder": item.get("placeholder"),
                "multiline": bool(item.get("multiline")),
                "options": options,
            }
        )

    return {
        "template_id": str(tm.raw.get("templateId") or tm.template_file),
        "template_file": tm.template_file,
        "fields": out_fields,
    }


def load_template(template_map: TemplateMap) -> Any:
    workbook_path = template_workbook_path(template_map.template_file)
    if not workbook_path.is_file():
        raise TemplateExportError(f"Template workbook not found: {workbook_path}")
    return load_workbook(workbook_path)


def _active_sheet(wb: Any, sheet_name: str | None) -> Worksheet:
    if sheet_name and sheet_name in wb.sheetnames:
        return wb[sheet_name]
    return wb.active


def _set_cell(ws: Worksheet, coord: str, value: Any) -> None:
    cell = ws[coord]
    if getattr(cell, "__class__", None).__name__ == "MergedCell":
        for merged in ws.merged_cells.ranges:
            if coord in merged:
                top = ws.cell(merged.min_row, merged.min_col)
                top.value = value
                return
        return
    cell.value = value


def populate_header(
    ws: Worksheet,
    template_map: TemplateMap,
    *,
    project: str,
    location: str,
    cost_object: str,
    comments: str,
    export_date: date | None = None,
    requester_name: str = "",
) -> None:
    when = export_date or date.today()
    if cell := template_map.header_cell("project"):
        _set_cell(ws, cell, project.strip())
    if cell := template_map.header_cell("location"):
        _set_cell(ws, cell, location.strip())
    if cell := template_map.header_cell("costObject"):
        _set_cell(ws, cell, cost_object.strip())
    if cell := template_map.header_cell("comments"):
        _set_cell(ws, cell, comments.strip())
    if cell := template_map.header_cell("date"):
        _set_cell(ws, cell, when)
    if name := (requester_name or "").strip():
        if cell := template_map.header_cell("requesterName"):
            _set_cell(ws, cell, name)


def _copy_cell_style(src: Cell, tgt: Cell) -> None:
    if not src.has_style:
        return
    tgt.font = copy(src.font)
    tgt.border = copy(src.border)
    tgt.fill = copy(src.fill)
    tgt.number_format = copy(src.number_format)
    tgt.protection = copy(src.protection)
    tgt.alignment = copy(src.alignment)


def _copy_row_style(ws: Worksheet, src_row: int, tgt_row: int, max_col: int = 12) -> None:
    if src_row in ws.row_dimensions:
        h = ws.row_dimensions[src_row].height
        if h is not None:
            ws.row_dimensions[tgt_row].height = h
    for c in range(1, max_col + 1):
        _copy_cell_style(ws.cell(src_row, c), ws.cell(tgt_row, c))


def _template_data_row_capacity(template_map: TemplateMap) -> int:
    first = template_map.first_data_row
    return max(0, template_map.last_template_data_row - first + 1)


def _ensure_rows(ws: Worksheet, template_map: TemplateMap, needed_rows: int) -> tuple[int, int]:
    """Ensure worksheet has enough styled data rows; return (style source row, inserted row count)."""
    first = template_map.first_data_row
    style_row = min(template_map.last_template_data_row, first)
    existing = _template_data_row_capacity(template_map)
    if needed_rows <= existing:
        return style_row, 0

    extra = needed_rows - existing
    insert_at = template_map.last_template_data_row + 1
    ws.insert_rows(insert_at, amount=extra)
    for i in range(extra):
        tgt = insert_at + i
        _copy_row_style(ws, style_row, tgt)
    return style_row, extra


def _footer_column_range(columns: str, row: int) -> str:
    parts = [p.strip() for p in str(columns).split(":") if p.strip()]
    if len(parts) == 2:
        return f"{parts[0]}{row}:{parts[1]}{row}"
    col = parts[0] if parts else "C"
    return f"{col}{row}"


def _add_list_validation(ws: Worksheet, sqref: str, list_range: str) -> None:
    formula = list_range if list_range.startswith("=") else f"={list_range}"
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showInputMessage=True,
        showErrorMessage=True,
    )
    dv.add(sqref)
    ws.add_data_validation(dv)


def apply_template_data_validations(
    ws: Worksheet,
    template_map: TemplateMap,
    *,
    row_count: int,
    inserted_rows: int = 0,
) -> None:
    """Re-apply Excel list validations (openpyxl drops x14 extension validations on load)."""
    cfg = template_map.data_validation_config()
    if not cfg:
        return

    lists_sheet = str(cfg.get("listsSheet") or "Lists")
    if lists_sheet not in ws.parent.sheetnames:
        return

    first = template_map.first_data_row
    last_data_row = first + max(row_count, 1) - 1
    if inserted_rows <= 0:
        last_data_row = max(last_data_row, template_map.last_template_data_row)

    for item in cfg.get("lineItems") or []:
        if not isinstance(item, dict):
            continue
        col = str(item.get("column") or "").strip()
        list_range = str(item.get("listRange") or "").strip()
        if not col or not list_range:
            continue
        _add_list_validation(ws, f"{col}{first}:{col}{last_data_row}", list_range)

    for item in cfg.get("footer") or []:
        if not isinstance(item, dict):
            continue
        base_row = int(item.get("row") or 0)
        columns = str(item.get("columns") or "C:D")
        list_range = str(item.get("listRange") or "").strip()
        if base_row <= 0 or not list_range:
            continue
        row = base_row + inserted_rows
        _add_list_validation(ws, _footer_column_range(columns, row), list_range)


def populate_rows(
    ws: Worksheet,
    template_map: TemplateMap,
    rows: Sequence[Mapping[str, Any]],
) -> None:
    if not rows:
        raise TemplateExportError("No items to export")

    cols = template_map.populate_columns()
    first = template_map.first_data_row
    style_row, inserted_rows = _ensure_rows(ws, template_map, len(rows))

    line_col = (template_map.raw.get("columns") or {}).get("lineNumber", "A")

    for idx, item in enumerate(rows):
        row_num = first + idx
        if row_num > template_map.last_template_data_row and row_num != first:
            _copy_row_style(ws, style_row, row_num)

        reimb = item.get("reimbursable")
        if reimb is True or (isinstance(reimb, str) and reimb.strip().upper() in ("Y", "YES", "TRUE", "1")):
            reimb_text = "YES"
        else:
            reimb_text = "NO"

        values: dict[str, Any] = {
            "reimbursable": reimb_text,
            "vendor": item.get("vendor") or "",
            "vendorPart": item.get("vendor_part_number") or item.get("vendorPart") or "",
            "description": item.get("item_name") or item.get("description") or "",
            "quantity": item.get("quantity"),
            "unit": item.get("unit") or "",
        }

        if line_col:
            ws[f"{line_col}{row_num}"].value = idx + 1

        for key, col_letter in cols.items():
            if key not in values:
                continue
            val = values[key]
            ws[f"{col_letter}{row_num}"].value = val

    apply_template_data_validations(ws, template_map, row_count=len(rows), inserted_rows=inserted_rows)


def save_workbook(wb: Any) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_filename(pattern: str, *, project: str, export_date: date | None = None) -> str:
    when = (export_date or date.today()).isoformat()
    safe_project = re.sub(r"[^\w\-]+", "_", project.strip().upper())[:48] or "PROJECT"
    return (
        pattern.replace("{project}", safe_project)
        .replace("{date}", when)
        .replace("{PROJECT}", safe_project)
        .replace("{DATE}", when)
    )


class TemplateExportService:
    """Configuration-driven Excel export from disk templates."""

    def __init__(self, map_path: Path | None = None) -> None:
        self._map_path = map_path

    def load_template_map(self) -> TemplateMap:
        return load_template_map(self._map_path)

    def material_request_form(self) -> dict[str, Any]:
        return build_material_request_template_form(self.load_template_map())

    def generate_template_map(self, workbook_path: Path, *, overwrite: bool = False) -> TemplateMap:
        from app.services.template_analyzer import generate_template_map

        data = generate_template_map(workbook_path, output_path=self._map_path or template_map_path(), overwrite=overwrite)
        return TemplateMap(data)

    def analyze_template(self, workbook_path: Path) -> dict[str, Any]:
        from app.services.template_analyzer import analyze_template

        return analyze_template(workbook_path)

    def load_template(self, template_map: TemplateMap | None = None) -> tuple[Any, TemplateMap]:
        tm = template_map or self.load_template_map()
        return load_template(tm), tm

    def populate_header(
        self,
        ws: Worksheet,
        template_map: TemplateMap,
        *,
        project: str,
        location: str,
        cost_object: str,
        comments: str,
        export_date: date | None = None,
        requester_name: str = "",
    ) -> None:
        populate_header(
            ws,
            template_map,
            project=project,
            location=location,
            cost_object=cost_object,
            comments=comments,
            export_date=export_date,
            requester_name=requester_name,
        )

    def populate_rows(self, ws: Worksheet, template_map: TemplateMap, rows: Sequence[Mapping[str, Any]]) -> None:
        populate_rows(ws, template_map, rows)

    def save_workbook(self, wb: Any) -> bytes:
        return save_workbook(wb)

    def export(
        self,
        rows: Sequence[Mapping[str, Any]],
        *,
        project: str,
        location: str,
        cost_object: str = "",
        comments: str = "",
        export_date: date | None = None,
        requester_name: str = "",
    ) -> tuple[bytes, str]:
        tm = self.load_template_map()
        wb = load_template(tm)
        ws = _active_sheet(wb, tm.sheet_name)
        populate_header(
            ws,
            tm,
            project=project,
            location=location,
            cost_object=cost_object,
            comments=comments,
            export_date=export_date,
            requester_name=requester_name,
        )
        populate_rows(ws, tm, rows)
        pattern = str(tm.raw.get("fileNamePattern") or "MR_{project}_{date}.xlsx")
        filename = build_filename(pattern, project=project, export_date=export_date)
        return save_workbook(wb), filename
