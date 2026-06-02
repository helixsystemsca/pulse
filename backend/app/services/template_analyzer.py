"""Analyze Excel templates and generate template-map.json."""

from __future__ import annotations

import json
import re
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.services.template_export_paths import template_map_path, templates_dir

# Labels we search for in the first ~35 rows to infer header / data layout.
_HEADER_LABEL_PATTERNS: list[tuple[str, str]] = [
    (r"project", "project"),
    (r"job\s*description|location", "location"),
    (r"^date", "date"),
    (r"cost\s*object", "costObject"),
    (r"comment", "comments"),
    (r"reimbursable", "reimbursable"),
    (r"vendor\s*part", "vendorPart"),
    (r"^vendor$", "vendor"),
    (r"description", "description"),
    (r"^qty", "quantity"),
    (r"^unit$", "unit"),
]


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", str(s).strip().lower())


def _cell_label(ws, row: int, col: int) -> str | None:
    v = ws.cell(row, col).value
    if v is None:
        return None
    return _norm(str(v))


def _find_label_cell(ws, pattern: str, max_row: int = 35, max_col: int = 12) -> tuple[int, int] | None:
    rx = re.compile(pattern, re.I)
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            label = _cell_label(ws, r, c)
            if label and rx.search(label):
                return r, c
    return None


def _value_cell_right_or_below(ws, row: int, col: int) -> str | None:
    for dc in (1, 2):
        v = ws.cell(row, col + dc).value
        if v is not None and str(v).strip():
            return f"{get_column_letter(col + dc)}{row}"
    for dr in (1,):
        v = ws.cell(row + dr, col).value
        if v is not None and str(v).strip() and not re.match(r"^[a-z].*:$", _norm(str(v))):
            return f"{get_column_letter(col)}{row + dr}"
    return None


def _find_header_row(ws) -> int | None:
    for r in range(1, 40):
        labels = {_cell_label(ws, r, c) for c in range(1, 12)}
        labels.discard(None)
        if any(l and "reimbursable" in l for l in labels) and any(l and "description" in l for l in labels):
            return r
    return None


def _column_letters_from_header_row(ws, header_row: int) -> dict[str, str]:
    cols: dict[str, str] = {}
    for c in range(1, 15):
        label = _cell_label(ws, header_row, c)
        if not label:
            continue
        for pattern, key in _HEADER_LABEL_PATTERNS:
            if key in cols:
                continue
            if re.search(pattern, label, re.I):
                cols[key] = get_column_letter(c)
    return cols


def analyze_template(
    workbook_path: Path,
    *,
    sheet_name: str | None = None,
) -> dict[str, Any]:
    """Inspect workbook layout and return a template-map structure."""
    wb = load_workbook(workbook_path, data_only=False)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    fields: dict[str, Any] = {}
    project_pos = _find_label_cell(ws, r"project")
    if project_pos:
        vc = _value_cell_right_or_below(ws, *project_pos)
        if vc:
            fields["project"] = {"cell": vc, "labelCell": f"{get_column_letter(project_pos[1])}{project_pos[0]}"}

    loc_pos = _find_label_cell(ws, r"job\s*description|location")
    if loc_pos:
        vc = _value_cell_right_or_below(ws, *loc_pos)
        if vc:
            fields["location"] = {"cell": vc, "labelCell": f"{get_column_letter(loc_pos[1])}{loc_pos[0]}"}

    date_pos = _find_label_cell(ws, r"^date")
    if date_pos:
        vc = _value_cell_right_or_below(ws, *date_pos)
        if vc:
            fields["date"] = {"cell": vc, "labelCell": f"{get_column_letter(date_pos[1])}{date_pos[0]}"}

    cost_pos = _find_label_cell(ws, r"cost\s*object")
    if cost_pos:
        vc = _value_cell_right_or_below(ws, *cost_pos)
        if vc:
            fields["costObject"] = {"cell": vc, "labelCell": f"{get_column_letter(cost_pos[1])}{cost_pos[0]}"}

    comment_pos = _find_label_cell(ws, r"comment")
    if comment_pos:
        vc = _value_cell_right_or_below(ws, *comment_pos) or f"{get_column_letter(comment_pos[1])}{comment_pos[0] + 1}"
        fields["comments"] = {"cell": vc, "labelCell": f"{get_column_letter(comment_pos[1])}{comment_pos[0]}"}

    header_row = _find_header_row(ws)
    cols = _column_letters_from_header_row(ws, header_row) if header_row else {}
    first_data_row = (header_row + 1) if header_row else 8

    # Infer last row with line numbers in column A
    last_data = first_data_row
    if header_row:
        for r in range(first_data_row, min(ws.max_row, 60) + 1):
            a = ws.cell(r, 1).value
            if a is not None and str(a).strip().isdigit():
                last_data = r

    populate = {k: cols[k] for k in ("reimbursable", "vendor", "vendorPart", "description", "quantity", "unit") if k in cols}

    return {
        "templateId": workbook_path.stem,
        "templateFile": workbook_path.name,
        "sheetName": ws.title,
        "version": 1,
        "discoveredAt": date.today().isoformat(),
        "fields": fields,
        "projectCell": fields.get("project", {}).get("cell"),
        "locationCell": fields.get("location", {}).get("cell"),
        "dateCell": fields.get("date", {}).get("cell"),
        "costObjectCell": fields.get("costObject", {}).get("cell"),
        "commentsCell": fields.get("comments", {}).get("cell"),
        "headerRow": header_row,
        "firstDataRow": first_data_row,
        "lastTemplateDataRow": last_data,
        "columns": cols,
        "populateColumns": populate,
        "fileNamePattern": "MR_{project}_{date}.xlsx",
    }


def generate_template_map(
    workbook_path: Path | None = None,
    *,
    output_path: Path | None = None,
    overwrite: bool = False,
) -> dict[str, Any]:
    """Write template-map.json from workbook analysis."""
    out = output_path or template_map_path()
    if out.exists() and not overwrite:
        with out.open(encoding="utf-8") as f:
            return json.load(f)

    path = workbook_path or (templates_dir() / "kent_material_request.xlsx")
    if not path.is_file():
        raise FileNotFoundError(f"Template workbook not found: {path}")

    mapping = analyze_template(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", encoding="utf-8") as f:
        json.dump(mapping, f, indent=2)
        f.write("\n")
    return mapping
